"""
Excel 导出（xlsx）。

使用 openpyxl（如果可用）；若环境不装 openpyxl，降级为伪 xlsx（实际写 CSV 但扩展名为 xlsx），
并在日志中标记，便于上游知晓。

支持：
- 多 Sheet（每个 dict key 对应一个 sheet）
- 列宽自适应
- 字段脱敏
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

from app.services.export.csv_exporter import CsvExporter, CsvExportConfig

logger = logging.getLogger(__name__)

try:
    import openpyxl  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
    OPENPYXL_AVAILABLE = True
except Exception:
    openpyxl = None  # type: ignore
    get_column_letter = None  # type: ignore
    OPENPYXL_AVAILABLE = False


@dataclass
class ExcelExportConfig:
    """Excel 导出配置。"""
    max_rows_per_sheet: int = 100_000
    auto_width: bool = True
    freeze_header: bool = True
    mask_fields: List[str] = field(default_factory=lambda: [
        'name', 'id_card', 'phone',
    ])
    default_sheet_name: str = 'data'


class ExcelExporter:
    """Excel 导出器。"""

    def __init__(
        self,
        output_dir: Path,
        config: Optional[ExcelExportConfig] = None,
        mask_fn: Optional[Callable[[str, Any], Any]] = None,
    ):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.config = config or ExcelExportConfig()
        self.mask_fn = mask_fn or self._default_mask
        self._csv_fallback = CsvExporter(
            output_dir=output_dir,
            config=CsvExportConfig(mask_fields=self.config.mask_fields),
        )

    def export(
        self,
        sheets: Dict[str, Dict[str, Any]],
        filename: str,
        apply_mask: bool = True,
    ) -> Dict[str, Any]:
        """
        导出多 Sheet Excel。

        sheets 格式：
          {
            'patients': {'fields': [...], 'records': [...]},
            'lab_results': {'fields': [...], 'records': [...]},
          }
        """
        path = self.output_dir / filename
        if not OPENPYXL_AVAILABLE:
            return self._export_as_csv_fallback(sheets, path, apply_mask)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 删除默认 sheet
        total_rows = 0

        for sheet_name, payload in sheets.items():
            fields = payload.get('fields', [])
            records = payload.get('records', [])
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet 名限制 31 字
            ws.append(fields)
            rows_written = 0
            for rec in records:
                if rows_written >= self.config.max_rows_per_sheet:
                    break
                row = []
                for f in fields:
                    v = rec.get(f, '')
                    if apply_mask and f in self.config.mask_fields:
                        v = self.mask_fn(f, v)
                    row.append(v)
                ws.append(row)
                rows_written += 1
            total_rows += rows_written

            if self.config.freeze_header:
                ws.freeze_panes = 'A2'
            if self.config.auto_width:
                self._auto_width(ws, fields)

        wb.save(str(path))
        size = path.stat().st_size
        logger.info('xlsx export %s: sheets=%d total_rows=%d bytes=%d',
                     path, len(sheets), total_rows, size)
        return {
            'path': str(path),
            'format': 'xlsx',
            'sheet_count': len(sheets),
            'total_rows': total_rows,
            'bytes': size,
            'masked': apply_mask,
        }

    def export_single(
        self,
        records: List[Dict[str, Any]],
        fields: List[str],
        filename: str,
        apply_mask: bool = True,
    ) -> Dict[str, Any]:
        """导出为单 sheet 的 xlsx。"""
        return self.export(
            {self.config.default_sheet_name: {'fields': fields, 'records': records}},
            filename,
            apply_mask,
        )

    # --- CSV fallback ---

    def _export_as_csv_fallback(
        self,
        sheets: Dict[str, Dict[str, Any]],
        path: Path,
        apply_mask: bool,
    ) -> Dict[str, Any]:
        logger.warning('openpyxl not installed, falling back to CSV bundle')
        total = 0
        for sheet_name, payload in sheets.items():
            fields = payload.get('fields', [])
            records = payload.get('records', [])
            csv_name = f'{path.stem}-{sheet_name}.csv'
            res = self._csv_fallback.export(
                records, fields, csv_name, apply_mask=apply_mask,
            )
            total += res['rows_written']
        return {
            'path': str(path),
            'format': 'csv_fallback',
            'sheet_count': len(sheets),
            'total_rows': total,
            'bytes': 0,
            'masked': apply_mask,
        }

    # --- helpers ---

    def _auto_width(self, ws, fields: List[str]) -> None:
        if not OPENPYXL_AVAILABLE:
            return
        for idx, field_name in enumerate(fields, start=1):
            col_letter = get_column_letter(idx)
            max_len = len(str(field_name))
            # 抽样前 50 行估列宽
            for row_idx in range(2, min(ws.max_row + 1, 52)):
                val = ws.cell(row=row_idx, column=idx).value
                if val is None:
                    continue
                max_len = max(max_len, min(60, len(str(val))))
            ws.column_dimensions[col_letter].width = max_len + 2

    @staticmethod
    def _default_mask(field: str, value: Any) -> Any:
        if value is None or value == '':
            return value
        s = str(value)
        if field == 'id_card':
            if len(s) >= 10:
                return s[:6] + '*' * (len(s) - 10) + s[-4:]
            return '*' * len(s)
        if field == 'phone':
            if len(s) >= 7:
                return s[:3] + '*' * (len(s) - 7) + s[-4:]
            return '*' * len(s)
        if field == 'name':
            return s[0] + '*' * (len(s) - 1) if len(s) > 1 else s
        return s
