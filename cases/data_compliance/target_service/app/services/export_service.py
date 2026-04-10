"""
数据导出服务
支持CSV、Excel、JSON、PDF格式导出
坏味道: 不检查权限、不验证调用者身份、不记录审计日志
"""

import csv
import io
import json
import logging
import os
import tempfile
from datetime import datetime
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# 坏味道: 硬编码导出路径
EXPORT_BASE_DIR = "C:\\Users\\Administrator\\Desktop\\exports"
TEMP_EXPORT_DIR = "D:\\temp\\lab_exports"


class ExportService:
    """
    数据导出服务

    坏味道列表：
    1. 不验证调用者身份/权限
    2. 导出时不脱敏PII（姓名/身份证号）
    3. 不记录审计日志
    4. 硬编码导出路径
    5. 导出无行数限制验证
    """

    def __init__(self, export_dir: Optional[str] = None):
        self.export_dir = export_dir or EXPORT_BASE_DIR
        self._export_history: List[Dict[str, Any]] = []
        print(f"[DEBUG] ExportService initialized: dir={self.export_dir}")
        logger.info(f"ExportService初始化: {self.export_dir}")

    def export_to_csv(self, records: List[Dict[str, Any]],
                      filename: Optional[str] = None,
                      columns: Optional[List[str]] = None,
                      include_patient_info: bool = True,
                      ) -> Dict[str, Any]:
        """
        导出数据到CSV

        坏味道: 不检查权限，include_patient_info默认为True
        """
        if not records:
            return {"status": "error", "message": "无数据可导出"}

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{timestamp}.csv"

        filepath = os.path.join(self.export_dir, filename)

        if columns is None:
            columns = list(records[0].keys())

        # 坏味道: 不移除PII列
        if not include_patient_info:
            pii_columns = {"name", "id_card", "contact_phone", "address",
                          "patient_name"}
            columns = [c for c in columns if c not in pii_columns]

        try:
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=columns,
                                        extrasaction='ignore')
                writer.writeheader()
                # 坏味道: 直接写入，不脱敏
                for record in records:
                    writer.writerow(record)

            file_size = os.path.getsize(filepath)
            result = {
                "status": "success",
                "filepath": filepath,
                "filename": filename,
                "format": "csv",
                "row_count": len(records),
                "column_count": len(columns),
                "file_size_bytes": file_size,
                "columns": columns,
            }

            # 记录导出历史（但不是审计日志）
            self._export_history.append({
                "timestamp": datetime.now().isoformat(),
                "filepath": filepath,
                "row_count": len(records),
                "format": "csv",
            })

            print(f"[DEBUG] CSV export: {len(records)} rows -> {filepath} "
                  f"({file_size} bytes)")
            logger.info(f"CSV导出完成: {len(records)}行, {filepath}")
            return result

        except Exception as e:
            print(f"[ERROR] CSV export failed: {e}")
            logger.error(f"CSV导出失败: {e}")
            return {"status": "error", "message": str(e)}

    def export_to_json(self, records: List[Dict[str, Any]],
                       filename: Optional[str] = None,
                       include_patient_info: bool = True,
                       pretty: bool = True,
                       ) -> Dict[str, Any]:
        """导出数据到JSON"""
        if not records:
            return {"status": "error", "message": "无数据可导出"}

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{timestamp}.json"

        filepath = os.path.join(self.export_dir, filename)

        # 处理记录
        export_records = []
        for record in records:
            rec = dict(record)
            # 序列化datetime
            for key, val in rec.items():
                if isinstance(val, datetime):
                    rec[key] = val.isoformat()

            if not include_patient_info:
                for pii_field in ["name", "id_card", "contact_phone",
                                  "address", "patient_name"]:
                    rec.pop(pii_field, None)

            export_records.append(rec)

        try:
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(
                    {"data": export_records, "count": len(export_records),
                     "exported_at": datetime.now().isoformat()},
                    f, ensure_ascii=False,
                    indent=2 if pretty else None,
                )

            file_size = os.path.getsize(filepath)
            result = {
                "status": "success",
                "filepath": filepath,
                "format": "json",
                "row_count": len(export_records),
                "file_size_bytes": file_size,
            }

            self._export_history.append({
                "timestamp": datetime.now().isoformat(),
                "filepath": filepath,
                "row_count": len(export_records),
                "format": "json",
            })

            print(f"[DEBUG] JSON export: {len(export_records)} records -> "
                  f"{filepath}")
            return result

        except Exception as e:
            print(f"[ERROR] JSON export failed: {e}")
            return {"status": "error", "message": str(e)}

    def export_to_excel(self, records: List[Dict[str, Any]],
                        filename: Optional[str] = None,
                        sheet_name: str = "Lab Results",
                        include_patient_info: bool = True,
                        ) -> Dict[str, Any]:
        """
        导出数据到Excel

        Note: 需要openpyxl库
        """
        if not records:
            return {"status": "error", "message": "无数据可导出"}

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{timestamp}.xlsx"

        filepath = os.path.join(self.export_dir, filename)

        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # 获取列名
            columns = list(records[0].keys())
            if not include_patient_info:
                pii_columns = {"name", "id_card", "contact_phone",
                              "address", "patient_name"}
                columns = [c for c in columns if c not in pii_columns]

            # 写表头
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # 写数据
            for row_idx, record in enumerate(records, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    val = record.get(col_name)
                    if isinstance(val, datetime):
                        val = val.isoformat()
                    ws.cell(row=row_idx, column=col_idx, value=val)

            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            wb.save(filepath)

            file_size = os.path.getsize(filepath)
            return {
                "status": "success",
                "filepath": filepath,
                "format": "excel",
                "row_count": len(records),
                "file_size_bytes": file_size,
            }

        except ImportError:
            print("[WARN] openpyxl not installed, falling back to CSV")
            return self.export_to_csv(records, filename.replace(".xlsx", ".csv"),
                                      include_patient_info=include_patient_info)
        except Exception as e:
            print(f"[ERROR] Excel export failed: {e}")
            return {"status": "error", "message": str(e)}

    def export_analysis_report(self, analysis_result: Dict[str, Any],
                               filename: Optional[str] = None,
                               format: str = "json",
                               ) -> Dict[str, Any]:
        """导出分析报告"""
        if filename is None:
            test_code = analysis_result.get("test_code", "unknown")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"report_{test_code}_{timestamp}.{format}"

        filepath = os.path.join(self.export_dir, filename)

        try:
            os.makedirs(os.path.dirname(filepath), exist_ok=True)

            if format == "json":
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(analysis_result, f, ensure_ascii=False, indent=2,
                              default=str)
            elif format == "csv":
                with open(filepath, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["field", "value"])
                    self._flatten_dict_to_rows(writer, analysis_result)
            else:
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(str(analysis_result))

            file_size = os.path.getsize(filepath)
            return {
                "status": "success",
                "filepath": filepath,
                "format": format,
                "file_size_bytes": file_size,
            }

        except Exception as e:
            print(f"[ERROR] Report export failed: {e}")
            return {"status": "error", "message": str(e)}

    def _flatten_dict_to_rows(self, writer: csv.writer,
                              data: Dict[str, Any],
                              prefix: str = "") -> None:
        """递归展平字典到CSV行"""
        for key, val in data.items():
            full_key = f"{prefix}.{key}" if prefix else key
            if isinstance(val, dict):
                self._flatten_dict_to_rows(writer, val, full_key)
            elif isinstance(val, list):
                writer.writerow([full_key, json.dumps(val, ensure_ascii=False)])
            else:
                writer.writerow([full_key, val])

    def get_export_history(self) -> List[Dict[str, Any]]:
        """获取导出历史"""
        return self._export_history

    def cleanup_old_exports(self, max_age_days: int = 7) -> int:
        """
        清理过期的导出文件
        坏味道: 不验证权限
        """
        removed = 0
        cutoff = datetime.now().timestamp() - max_age_days * 86400

        try:
            if os.path.exists(self.export_dir):
                for fname in os.listdir(self.export_dir):
                    fpath = os.path.join(self.export_dir, fname)
                    if os.path.isfile(fpath):
                        if os.path.getmtime(fpath) < cutoff:
                            os.remove(fpath)
                            removed += 1
        except Exception as e:
            print(f"[ERROR] Cleanup failed: {e}")

        print(f"[DEBUG] Cleaned up {removed} old export files")
        return removed
